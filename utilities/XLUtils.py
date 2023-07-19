# see->  python code with notes(Desktop)\\Ch.Data Driven Testing using Excel
#..before coding below.

import openpyxl

def getRowCount(file_path, sheetName):
    workBook = openpyxl.load_workbook(file_path)
    workSheet = workBook[sheetName]
    x = workSheet.max_row
    return x

def getColumnCount(file_path, sheetName):
    workBook = openpyxl.load_workbook(file_path)
    workSheet = workBook[sheetName]
    y = workSheet.max_column
    return y

def readData(file_path,sheetName,rowNum,columnNum):
    workBook = openpyxl.load_workbook(file_path)
    workSheet = workBook[sheetName]
    z = workSheet.cell(row= rowNum, column= columnNum).value
    return z

# keep yr excel sheet closed before running below Fn
def writeData(file_path,sheetName,rowNum,columnNum,data):
    workBook = openpyxl.load_workbook(file_path)
    workSheet = workBook[sheetName]
    workSheet.cell(row= rowNum,column= columnNum).value= data
    workSheet.save(file_path)

